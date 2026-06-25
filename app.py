import streamlit as st
import pandas as pd
import re
import io
import uuid
import json
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from yookassa import Configuration, Payment

st.set_page_config(page_title="Подбор вузов", layout="wide")

OBL = {
    "Русский язык": 0, "Математика": 1, "Обществознание": 2,
    "История": 3, "Иностранный язык": 4, "Биология": 5,
    "Химия": 6, "Физика": 7, "Информатика": 8,
    "География": 9, "Литература": 10, "ДВИ": 11,
}
VYB = {
    "Математика": 12, "Обществознание": 13, "История": 14,
    "Иностранный язык": 15, "Биология": 16, "Химия": 17,
    "Физика": 18, "Информатика": 19, "География": 20, "Литература": 21,
}

@st.cache_data
def load_data():
    # Канонический список названий колонок задаётся явно по позиции.
    # Это защищает от расхождений между листами Excel (например,
    # "Вуз" на одном листе и "ВУЗ" на другом, или лишний пустой столбец) —
    # такие расхождения ломают pd.concat, который объединяет по именам колонок,
    # а не по позиции, и приводят к "перекосу" данных между листами.
    CANONICAL_COLS = [
        "Русский язык", "Математика", "Обществознание", "История", "Иностранный язык",
        "Биология", "Химия", "Физика", "Информатика", "География", "Литература", "ДВИ",
        "Математика.1", "Обществознание.1", "История.1", "Иностранный язык.1",
        "Биология.1", "Химия.1", "Физика.1", "Информатика.1", "География.1", "Литература.1",
        "Город", "Вуз", "Факультет (институт/школа)",
        "Код и специальность \n(направление подготовки)",
        "Образовательная программа (профиль)",
        "Количество бюджетных мест 2026 (Всего)",
        "Проходной балл бюджет 2025", "Средний балл бюджет 2025",
        "Число зачисленных 2025", "Число зачисленных БВИ",
        "ГТО Золото", "ГТО Серебро", "ГТО Бронза", "Аттестат с отличием",
        "Все индивидуальные достижения, за которые вуз добавляет баллы (ссылка)",
        "Стоимость обучения в год, тыс \n(собираем после 1 июня)",
        "Комментарий",
    ]
    sheets = {}
    for sheet in ["Москва", "Питер", "Регионы"]:
        df = pd.read_excel("База вузов 2026.xlsx", sheet_name=sheet, header=1)
        n = len(CANONICAL_COLS)
        if len(df.columns) < n:
            st.error(f"На листе '{sheet}' меньше колонок ({len(df.columns)}), чем ожидается ({n}). Проверьте структуру файла.")
        df = df.iloc[:, :n].copy()
        df.columns = CANONICAL_COLS
        df["_лист"] = sheet
        sheets[sheet] = df
    full = pd.concat(sheets.values(), ignore_index=True)
    full.iloc[:, 25] = full.iloc[:, 25].astype(str).str.strip()
    return full

def get_city_group(city_val):
    s = str(city_val).strip()
    if any(x in s for x in ["Москва", "МО", "Московская"]):
        return "Москва и Московская область"
    if any(x in s for x in ["Петербург", "Ленинград", "Пушкин", "Гатчина"]):
        return "Санкт-Петербург и Ленинградская область"
    return s

@st.cache_data
def get_city_options(df):
    cities = df.iloc[:, 22].dropna().unique()
    groups = sorted(set(get_city_group(c) for c in cities))
    priority = ["Москва и Московская область", "Санкт-Петербург и Ленинградская область"]
    return priority + [g for g in groups if g not in priority]

@st.cache_data
def get_vuz_by_city(df, city_group):
    mask = df.iloc[:, 22].apply(lambda x: get_city_group(x) == city_group)
    vuzы = df[mask].iloc[:, 23].dropna().unique()
    return sorted(set(str(v).strip() for v in vuzы if str(v).strip() not in ("", "nan")))

@st.cache_data
def get_all_codes(df):
    codes = df.iloc[:, 25].dropna().unique()
    return sorted(set(str(c).strip() for c in codes if str(c).strip() not in ("", "nan")))

def clean_str(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s

def to_num(val):
    s = str(val).strip()
    if s.upper() == "NEW": return "NEW"
    if s == "-": return "-"
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except:
        return None

def cell_has_value(row, col_idx):
    val = row.iloc[col_idx]
    if pd.isna(val): return False
    s = str(val).strip()
    if s in ('', 'nan'): return False
    try: return float(s) > 0
    except: return True

def cell_value(row, col_idx):
    try: return float(row.iloc[col_idx])
    except: return 0

def subject_passes(row, col_idx, student_score):
    if not cell_has_value(row, col_idx): return False
    threshold = cell_value(row, col_idx)
    if threshold == 0: return True
    return student_score >= threshold

def has_budget_places(val):
    s = str(val).strip()
    if s in ('-', 'nan', ''): return False
    m = re.match(r'(\d+)\s*\((\d+)\)', s)
    if m and int(m.group(2)) == 0: return True  # ИЗМЕНЕНО: теперь пропускаем X(0) дальше
    try: return float(s) > 0
    except: return True

def is_quota_bvi(val):
    """X(0) — места выделены только под квоты/БВИ, общего конкурса нет"""
    s = str(val).strip()
    m = re.match(r'(\d+)\s*\((\d+)\)', s)
    return bool(m and int(m.group(2)) == 0)

def is_no_competition(row):
    """AC=1, AD=1, AE=0 — в прошлом году общего конкурса не было"""
    try:
        return float(row.iloc[28]) == 1 and float(row.iloc[29]) == 1 and float(row.iloc[30]) == 0
    except (ValueError, TypeError):
        return False

def calc_achievements(row, gto, attestat):
    total = 0
    if gto == "Золото": total += cell_value(row, 32)
    elif gto == "Серебро": total += cell_value(row, 33)
    elif gto == "Бронза": total += cell_value(row, 34)
    if attestat: total += cell_value(row, 35)
    return min(total, 10)

def calc_student_score(row, subjects):
    score = subjects.get("Русский язык", 0)
    obl_set = set()
    for subj, col in OBL.items():
        if subj in ("Русский язык", "ДВИ"): continue
        if cell_has_value(row, col):
            score += subjects.get(subj, 0)
            obl_set.add(subj)
    best_vyb = 0
    for subj, col in VYB.items():
        if subj in obl_set: continue
        if cell_has_value(row, col):
            s = subjects.get(subj, 0)
            if s > best_vyb: best_vyb = s
    score += best_vyb
    return score

def check_row(row, subjects):
    rus = subjects.get("Русский язык", 0)
    if not cell_has_value(row, OBL["Русский язык"]): return None
    if rus < cell_value(row, OBL["Русский язык"]): return None
    obl_required = [(s, c) for s, c in OBL.items()
                    if s not in ("Русский язык", "ДВИ") and cell_has_value(row, c)]
    for subj, col in obl_required:
        if not subject_passes(row, col, subjects.get(subj, 0)): return None
    dvi_required = cell_has_value(row, OBL["ДВИ"])
    obl_set = {s for s, _ in obl_required}
    vyb_required = [(s, c) for s, c in VYB.items()
                    if s not in obl_set and cell_has_value(row, c)]
    if vyb_required:
        if not any(subjects.get(s, 0) > 0 and subject_passes(row, c, subjects.get(s, 0))
                   for s, c in vyb_required): return None
    return "with_dvi" if dvi_required else "no_dvi"

def is_valid_score(v):
    """True если значение — число больше 1 (не пустое, не прочерк, не 0 или 1)"""
    try:
        return float(v) > 1
    except:
        return False

def get_chance(student_score, pb, sb):
    pb_ok = is_valid_score(pb)
    sb_ok = is_valid_score(sb)
    if not pb_ok and not sb_ok:
        return "new"
    if pb_ok and sb_ok:
        pb_f, sb_f = float(pb), float(sb)
        if student_score >= sb_f + 10: return "podstrahovka"
        if student_score >= sb_f: return "realistic"
        if student_score > pb_f + 5: return "probable"
        if student_score >= pb_f - 15: return "risky"
        return "unlikely"
    # Только ПБ доступен — СБ отсутствует (часть региональных вузов)
    pb_f = float(pb)
    if student_score >= pb_f + 25: return "podstrahovka"
    if student_score >= pb_f + 10: return "realistic"
    if student_score > pb_f + 5: return "probable"
    if student_score >= pb_f - 15: return "risky"
    return "unlikely"
VUZ_RATING = {
    "МГУ им. Ломоносова": 1,
    "МГТУ им. Баумана": 2,
    "МФТИ": 3,
    "СПБГУ": 4,
    "МИФИ": 5,
    "ВШЭ (Высшая школа экономики)": 6,
    "МГИМО": 7,
    "РАНХиГС (Российская академия народного хозяйства": 8,
    "Политех им. Петра Великого": 9,
    "Финансовый университет": 10,
    "МГМУ им. Сеченова": 11,
    "УрФУ": 12,
    "ТПУ": 13,
    "РУДН": 14,
    "ИТМО": 15,
    "НГУ (Новосибирский государственный университет)": 16,
    "МИСиС": 17,
    "РЭУ им. Г.В. Плеханова": 18,
    "ТГУ (Томский": 19,
    "КФУ (Казанский (Приволжский)": 20,
    "РНИМУ им. Пирогова": 21,
    "МАИ": 22,
    "СФУ (Сибирский федеральный": 23,
    "МЭИ": 24,
    "ДВФУ": 25,
    "МГЮА им. Кутафина": 26,
    "РГУНиГ им. Губкина": 27,
    "Горный университет": 28,
    "ПСПБГМУ им. Павлова": 29,
    "МГСУ": 30,
    "ЮФУ": 31,
    "МПГУ (Московский педагогический": 32,
    "МГЛУ": 33,
    "ЛЭТИ": 34,
    "МСХА им. Тимирязева": 35,
    "ННГУ им. Н. И. Лобачевского": 36,
    "МИРЭА": 37,
    "ВАВТ": 38,
    "УГНТУ": 39,
    "РХТУ им. Менделеева": 40,
    "РГПУ им. Герцена": 41,
    "ЮУрГУ": 42,
    "БелГУ (Белгородский государственный национальный": 43,
    "МГПУ (Московский городской": 44,
    "Самарский университет (Самарский национальный": 45,
    "СПБГЭУ": 46,
    "НГТУ (Новосибирский государственный технический": 47,
    "СтГАУ": 48,
    "СПБГПМУ": 49,
    "РГГУ": 50,
    "Казанский КГМУ": 51,
    "СамГМУ": 52,
    "ТюмГУ": 53,
    "НМИЦ им. Алмазова": 55,
    "СКФУ": 56,
    "РязГМУ": 57,
    "ЮРГПУ (НПИ)": 58,
    "ДГТУ": 59,
    "МГТУ Станкин": 60,
    "СВФУ": 61,
    "ПИМУ": 62,
    "ТУСУР": 63,
    "БГМУ (Башкирский": 64,
    "СПБГАСУ": 65,
    "КубГАУ": 66,
    "ПНИПУ": 67,
    "БФУ имени И. Канта": 68,
    "КНИТУ-КАИ": 69,
    "КГМУ": 70,
    "СЗГМУ им. Мечникова": 71,
    "СибГМУ": 72,
    "ВолгГМУ": 73,
    "МАДИ": 74,
    "РУТ (МИИТ)": 75,
    "Дубна": 76,
    "КНИТУ (Казанский национальный исследовательский технологический": 77,
    "АлтГУ (Алтайский государственный университет)": 78,
    "Политех (Московский политехнический": 80,
    "СГУ (Саратовский национальный": 81,
    "КБГУ": 82,
    "ИРНИТУ": 83,
    "СПБГМТУ": 84,
    "ГГТУ": 85,
    "БГТУ им. В.Г. Шухова": 86,
    "СамГТУ": 87,
    "ГСГУ": 88,
    "КубГМУ": 89,
    "Саратовский ГМУ": 90,
    "ГУУ (Государственный университет управления)": 91,
    "МИЭТ": 92,
    "ВГМУ им. Н.Н. Бурденко": 93,
    "ВолгГТУ": 94,
    "ТИУ": 95,
    "ПГУПС": 96,
    "КрасГМУ": 97,
    "МТУСИ": 99,
    "УУНиТ": 100,
}

AREA_GROUPS = {
    "IT и программирование":            ["02", "09", "10"],
    "Математика и физика":              ["01", "03"],
    "Инженерия и технологии":           ["11", "12", "13", "14", "15", "16", "17", "22", "27", "28", "29"],
    "Химия и материалы":                ["04", "18"],
    "Архитектура, строительство и дизайн": ["07", "08", "54"],
    "Транспорт и авиация":              ["23", "24", "25", "26"],
    "Науки о земле и экология":         ["05", "20", "21"],
    "Биология":                         ["06"],
    "Медицина и фармация":              ["30", "31", "32", "33", "34"],
    "Сельское хозяйство и ветеринария": ["35", "36"],
    "Экономика и управление":           ["38"],
    "Юриспруденция":                    ["40"],
    "Психология и социология":          ["37", "39"],
    "Политология и медиа":              ["41", "42"],
    "Педагогика":                       ["44"],
    "Лингвистика и филология":          ["45"],
    "История и гуманитарные науки":     ["46", "47", "48"],
    "Сервис и туризм":                  ["43"],
    "Физическая культура и спорт":      ["49"],
    "Искусство и творчество":           ["50", "51", "52", "53", "55"],
}

def get_vuz_rating(vuz_name):
    """Возвращает позицию вуза в рейтинге или 999 если не найден"""
    for key, rank in VUZ_RATING.items():
        if key in vuz_name:
            return rank
    return 999
CHANCE_ORDER = {
    "probable": 0, "realistic": 1, "podstrahovka": 2,
    "risky": 3, "unlikely": 4,
    "quota_bvi": 5, "no_competition": 6,
    "new": 7, "no_dvi_score": 8,
}
CHANCE_LABEL = {
    "podstrahovka":    "🟢 Уверенно",
    "realistic":       "🔵 Реалистично",
    "probable":        "🟡 Вероятно",
    "risky":           "🔴 Рискованно",
    "unlikely":        "⚫ Маловероятно",
    "quota_bvi":        "🔹 Квоты и БВИ",
    "no_competition":   "◾ Общего конкурса не было",
    "new":             "⬜ Нет данных",
    "no_dvi_score":    "⬜ Нет оценки — не указан балл за ДВИ",
}
PRIORITY_LABEL = {
    "podstrahovka":    "3–5",
    "realistic":       "2–3",
    "probable":        "1–2",
    "risky":           "1*",
    "unlikely":        "—",
    "quota_bvi":        "—",
    "no_competition":   "1*",
    "new":             "1*",
    "no_dvi_score":    "—",
}

def build_result_row(row, subjects, gto, attestat, dvi_score=None):
    pb, sb = row.iloc[28], row.iloc[29]
    student_score = calc_student_score(row, subjects)
    achievements = calc_achievements(row, gto, attestat)
    total_score = student_score + achievements
    dvi_required = cell_has_value(row, OBL["ДВИ"])

    if dvi_required and dvi_score and dvi_score > 0:
        total_score += dvi_score

    if is_quota_bvi(row.iloc[27]):
        chance = "quota_bvi"
    elif is_no_competition(row):
        chance = "no_competition"
    elif dvi_required:
        if dvi_score and dvi_score > 0:
            chance = get_chance(total_score, pb, sb)
        else:
            chance = "no_dvi_score"
    else:
        chance = get_chance(total_score, pb, sb)
    return {
        "Город":               clean_str(row.iloc[22]),
        "Вуз":                 clean_str(row.iloc[23]),
        "Факультет":           clean_str(row.iloc[24]),
        "Код и специальность": clean_str(row.iloc[25]),
        "Профиль":             clean_str(row.iloc[26]),
        "Мест":                to_num(row.iloc[27]),
        "Проходной балл":      to_num(pb),
        "Средний балл":        to_num(sb),
        "Ваш балл (ЕГЭ)":     student_score,
        "Достижения":          achievements if achievements > 0 else "",
        "Конкурсный балл":     total_score,
        "Шансы":               chance,
        "Рек. приоритет":      PRIORITY_LABEL[chance],
        "ГТО золото":          to_num(row.iloc[32]),
        "ГТО серебро":         to_num(row.iloc[33]),
        "ГТО бронза":          to_num(row.iloc[34]),
        "Аттестат":            to_num(row.iloc[35]),
        "Стоимость обучения (Москва и СПб), тыс руб": clean_str(row.iloc[37]) if clean_str(row.iloc[37]) else "—",
    }

def expand_code_set(selected_codes, df):
    """Расширяет набор кодов с учётом трёхуровневой структуры Минобрнауки:
    ХХ.00.00 = УГСН (верхний уровень)
    ХХ.ХХ.00 = групповой уровень
    ХХ.ХХ.ХХ = конкретная специальность
    """
    all_db_codes = set(
        str(c).strip() for c in df.iloc[:, 25].tolist()
        if pd.notna(c) and str(c).strip() and str(c).strip().lower() != 'nan'
    )
    expanded = set(selected_codes)
    for code in selected_codes:
        if not code or str(code).strip().lower() == 'nan':
            continue
        num_part = code.split(' ')[0]
        parts = num_part.split('.')
        if len(parts) != 3:
            continue
        a, b, c = parts[0], parts[1], parts[2]
        if b == '00' and c == '00':
            # ХХ.00.00 → добавляем всё что начинается с ХХ.
            for db_code in all_db_codes:
                if db_code.split(' ')[0].startswith(a + '.'):
                    expanded.add(db_code)
        elif c == '00':
            # ХХ.ХХ.00 → добавляем все ХХ.ХХ.хх + родительский ХХ.00.00
            prefix = a + '.' + b
            for db_code in all_db_codes:
                db_num = db_code.split(' ')[0]
                if db_num.startswith(prefix + '.') or db_num == a + '.00.00':
                    expanded.add(db_code)
        else:
            # ХХ.ХХ.ХХ → добавляем родительские ХХ.ХХ.00 и ХХ.00.00
            for db_code in all_db_codes:
                db_num = db_code.split(' ')[0]
                if db_num == a + '.' + b + '.00' or db_num == a + '.00.00':
                    expanded.add(db_code)
    return expanded

def filter_rows_flow2(df, subjects, show_dvi, selected_city_groups, gto, attestat, dvi_score=None):
    results = []
    for _, row in df.iterrows():
        city_raw = str(row.iloc[22]).strip()
        if get_city_group(city_raw) not in selected_city_groups: continue
        if not has_budget_places(row.iloc[27]): continue
        status = check_row(row, subjects)
        if status is None: continue
        if status == "with_dvi" and not show_dvi: continue
        results.append(build_result_row(row, subjects, gto, attestat, dvi_score))
    return pd.DataFrame(results)

def filter_rows_flow1(df, subjects, selected_vuz, selected_codes, gto, attestat, selected_cities=None, dvi_score=None):
    expanded_codes = expand_code_set(selected_codes, df) if selected_codes else set()
    results = []
    for _, row in df.iterrows():
        city_raw = str(row.iloc[22]).strip()
        city_group = get_city_group(city_raw)
        vuz = clean_str(row.iloc[23])
        code = clean_str(row.iloc[25])
        if selected_cities and city_group not in selected_cities: continue
        if selected_vuz and vuz not in selected_vuz: continue
        if expanded_codes and code not in expanded_codes: continue
        if not selected_vuz and not expanded_codes: continue
        if not has_budget_places(row.iloc[27]): continue
        status = check_row(row, subjects)
        if status is None: continue
        results.append(build_result_row(row, subjects, gto, attestat, dvi_score))
    return pd.DataFrame(results)

def count_slots(selected_codes):
    slots = set()
    for code in selected_codes:
        parts = code.split(' ')[0]
        if len(parts) >= 7:
            prefix = parts[:5]
            suffix = parts[5:]
            if suffix == '.00':
                slots.add(prefix + '.00')
            else:
                has_multi = any(c.split(' ')[0] == prefix + '.00' for c in selected_codes)
                slots.add(prefix + '.00' if has_multi else code)
        else:
            slots.add(code)
    return len(slots)
# Глобальное хранилище платежей (живёт пока приложение запущено)
if "payment_store" not in st.session_state:
    st.session_state["payment_store"] = {}

import gspread
from google.oauth2.service_account import Credentials

def get_sheets_client():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=["https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive"]
    )
    return gspread.authorize(creds)

def save_payment_data(order_id, result_df, search_params, user_email, flow, payment_id=None):
    # Сохраняем в session_state
    if "payment_store" not in st.session_state:
        st.session_state["payment_store"] = {}
    st.session_state["payment_store"][order_id] = {
        "payment_id": str(payment_id) if payment_id else "",
        "user_email": user_email,
        "flow": flow,
        "search_params": {k: str(v) for k, v in search_params.items()},
        "result": result_df.to_dict(),
    }
    # Сохраняем в /tmp (доступно на Streamlit Cloud, страховка для return_url-флоу)
    try:
        filename = f"/tmp/payment_{order_id}.json"
        data = {
            "payment_id": str(payment_id) if payment_id else "",
            "user_email": user_email,
            "flow": flow,
            "search_params": {k: str(v) for k, v in search_params.items()},
            "result": result_df.to_dict(),
        }
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        st.warning(f"Не удалось сохранить в файл: {e}")
    # Сохраняем в Google Sheets — включая полный JSON с результатами (сжатый),
    # чтобы внешний webhook-сервис (Render) тоже мог прочитать данные
    # и отправить письмо независимо от Streamlit-приложения.
    try:
        import gzip
        import base64
        client = get_sheets_client()
        sheet = client.open_by_key(st.secrets["SHEETS_ID"]).sheet1
        # Оставляем только нужные колонки для письма
        cols_to_keep = [c for c in ["Город", "Вуз", "Факультет", "Код и специальность", "Профиль",
            "Мест", "Проходной балл", "Средний балл", "Ваш балл (ЕГЭ)", "Достижения",
            "Конкурсный балл", "Шансы", "Стоимость обучения (Москва и СПб), тыс руб"] if c in result_df.columns]
        result_df_slim = result_df[cols_to_keep].copy()
        # Для флоу 2 оставляем только вузы с минимум 3 подходящими вариантами
        if flow == 2 and "Вуз" in result_df_slim.columns:
            vuz_counts = result_df_slim.groupby("Вуз")["Код и специальность"].count()
            vuzы_ok = vuz_counts[vuz_counts >= 3].index
            result_df_slim = result_df_slim[result_df_slim["Вуз"].isin(vuzы_ok)]
        chance_priority = {
            "podstrahovka": 0, "realistic": 1, "probable": 2,
            "risky": 3, "new": 4, "quota_bvi": 5,
            "no_competition": 6, "unlikely": 7, "no_dvi_score": 8,
        }
        if "Шансы" in result_df_slim.columns and "Вуз" in result_df_slim.columns:
            result_df_slim["_chance_sort"] = result_df_slim["Шансы"].map(lambda x: chance_priority.get(x, 9))
            result_df_slim["_rating_sort"] = result_df_slim["Вуз"].map(lambda x: get_vuz_rating(x))
            if flow == 2:
                # Флоу 2: зеркалим логику экрана — топ-7 вузов, порог как в show_results
                good_zones_raw = {"podstrahovka", "realistic", "probable"}
                good_rows = result_df_slim[result_df_slim["Шансы"].isin(good_zones_raw)]
                vuz_good_count = good_rows.groupby("Вуз").size()
                vuz_with_3plus = vuz_good_count[vuz_good_count >= 3]
                min_good_options = 1 if len(vuz_with_3plus) < 3 else 3
                main_vuz = vuz_good_count[vuz_good_count >= min_good_options]
                # Сортируем: сначала рейтинговые с ≥4 вариантами, потом остальные
                rated = sorted(
                    [v for v in main_vuz.index if get_vuz_rating(v) < 999 and main_vuz[v] >= 4],
                    key=lambda v: get_vuz_rating(v)
                )
                unrated = [v for v in main_vuz.index if v not in rated]
                top_vuz = (rated + unrated)[:7]
                # Для каждого вуза берём до 5 уникальных кодов
                rows_out = []
                for vuz in top_vuz:
                    vuz_df = result_df_slim[result_df_slim["Вуз"] == vuz].copy()
                    vuz_df = vuz_df.sort_values("_chance_sort")
                    seen_codes = set()
                    for _, row in vuz_df.iterrows():
                        code_prefix = str(row["Код и специальность"]).split(" ")[0][:5]
                        if code_prefix not in seen_codes:
                            if len(seen_codes) >= 5:
                                continue
                            seen_codes.add(code_prefix)
                        rows_out.append(row)
                result_df_slim = pd.DataFrame(rows_out).reset_index(drop=True)
                result_df_slim = result_df_slim.sort_values(["Город", "Вуз", "_chance_sort"]).drop(columns=["_chance_sort", "_rating_sort"])
            else:
                # Флоу 1: лимита нет — пользователь сам выбрал ≤5 вузов
                result_df_slim = result_df_slim.sort_values(["Город", "Вуз", "_chance_sort"]).drop(columns=["_chance_sort", "_rating_sort"])
        else:
            pass  # оставляем result_df_slim как есть
        full_data = {
            "user_email": user_email,
            "flow": flow,
            "search_params": {k: str(v) for k, v in search_params.items()},
            "result": result_df_slim.to_dict(),
        }
        raw_json = json.dumps(full_data, ensure_ascii=False)
        compressed_b64 = base64.b64encode(gzip.compress(raw_json.encode("utf-8"))).decode("ascii")
        row = [
            order_id,
            str(payment_id) if payment_id else "",
            user_email,
            str(flow),
            datetime.now().isoformat(),
            "",  # статус отправки письма — заполняется webhook-сервисом ("sent")
            compressed_b64,
        ]
        all_values = sheet.get_all_values()
        next_row = len(all_values) + 1
        sheet.update(f"A{next_row}", [row], value_input_option="RAW")
    except Exception as e:
        import traceback
        st.warning(f"Не удалось сохранить в Google Sheets: {e}")
        st.warning(traceback.format_exc())
def get_email_sent_status(order_id):
    """Проверяет в Google Sheets, было ли уже отправлено письмо для этого order_id"""
    try:
        client = get_sheets_client()
        sheet = client.open_by_key(st.secrets["SHEETS_ID"]).sheet1
        cell = sheet.find(order_id, in_column=1)
        if cell:
            status = sheet.cell(cell.row, 6).value  # столбец F
            return status == "sent"
    except Exception:
        pass
    return False

def mark_email_sent(order_id):
    """Отмечает в Google Sheets, что письмо отправлено"""
    try:
        client = get_sheets_client()
        sheet = client.open_by_key(st.secrets["SHEETS_ID"]).sheet1
        cell = sheet.find(order_id, in_column=1)
        if cell:
            sheet.update_cell(cell.row, 6, "sent")  # столбец F
    except Exception:
        pass
def load_payment_data(order_id):
    # Сначала session_state
    if order_id in st.session_state.get("payment_store", {}):
        return st.session_state["payment_store"][order_id]
    # Потом /tmp файл
    try:
        filename = f"/tmp/payment_{order_id}.json"
        if os.path.exists(filename):
            with open(filename, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        st.warning(f"Не удалось загрузить из файла: {e}")
    return None

def check_payment_status(payment_id):
    """Проверяем статус платежа через API ЮКассы"""
    try:
        payment = Payment.find_one(payment_id)
        return payment.status == "succeeded"
    except:
        return False
def create_payment(amount, description, return_url, order_id=None):
    if order_id is None:
        order_id = str(uuid.uuid4())
    payment = Payment.create({
        "amount": {"value": str(amount), "currency": "RUB"},
        "confirmation": {"type": "redirect", "return_url": return_url},
        "capture": True,
        "description": description,
        "metadata": {"order_id": order_id}
    })
    return payment.confirmation.confirmation_url, payment.id

def show_disclaimers():
    with st.expander("📖 Как читать таблицу и расставлять приоритеты"):
        st.markdown("""
**Расшифровка шансов:**
- 🟢 **Уверенно** — ваш балл выше среднего на 10+. Чаще всего это подстраховочный вариант на который вы точно пройдёте. Рекомендуемый приоритет: **3–5+**
- 🔵 **Реалистично** — ваш балл выше среднего. Хорошие шансы. Рекомендуемый приоритет: **2–3**
- 🟡 **Вероятно** — ваш балл ниже среднего, но выше проходного на 5+. Шансы есть. Рекомендуемый приоритет: **1–2**
- 🔴 **Рискованно** — ваш балл близко к проходному (не более чем на 5 выше или 15 ниже). Ставьте приоритет 1 только если очень хочется и есть варианты с уверенным поступлением
- ⚫ **Маловероятно** — ваш балл ниже проходного на 15+. Шансы крайне малы
- 🔹 **Квоты и БВИ** — несмотря на наличие бюджетных мест, на данный момент они выделены только под квоты. Если желающих по квотам на эти специальности не будет, места перейдут олимпиадникам, а затем обычным поступающим на основном этапе. Это станет известно 3 августа
- ◾ **Общего конкурса не было** — в прошлом году зачисление на основном этапе не проводилось, были зачислены только абитуриенты по квотам и БВИ (олимпиадники)
- ⬜ **Нет данных** — новая специальность, статистики нет. Ставьте приоритет 1 только если очень хочется и есть варианты с уверенным поступлением

> 📊 **Про точность оценки шансов:** для части региональных вузов средний балл зачисленных недоступен — есть только проходной. В таких строках оценка шансов сделана только на основе проходного балла и может быть менее точной, чем для вузов где есть оба показателя.

**Про приоритеты:**
В одном вузе можно выбрать не более 5 кодов специальностей, но количество профилей не ограничено — поэтому приоритетов может быть больше пяти.

- 1–2 приоритет: амбициозные, но могут не сработать
- 3–4 приоритет: реалистичные, стабильные
- 4+ приоритет: уверенное зачисление
        """)
    st.warning("""
🔴 **Самое важное — согласие на зачисление**

Зачисление не происходит без подачи согласия. Его нужно подать до **12:00 (мск) 5 августа** в один из вузов — только один одновременно.

Эта таблица поможет выбрать куда подать заявление и документы, но участвовать в конкурсе вы будете только в одном вузе.
    """)
    st.markdown("""
---
**Следующий шаг** — отслеживание позиции в конкурсных списках на сайтах вузов или на Госуслугах. Не бойтесь раздутых списков — не все абитуриенты в них ваши реальные конкуренты.

**Актуальность данных.** Данные актуальны на приёмную кампанию 2026 года. Результаты основаны на статистике прошлого года и носят рекомендательный характер.

**Сроки подачи документов.** Документы в вузы нужно подать до **25 июля**. Срок для ДВИ короче — даты и формат уточняйте на сайте вуза.

**Индивидуальные достижения.** Если у вас есть достижения которые принимает вуз, но они не учтены в нашей таблице — прибавьте их самостоятельно. Помните: суммарно не более 10 баллов. Полный список на сайте вуза.

*Таблица только для поступающих на общих основаниях на основном этапе. Квотники и олимпиадники — вам полезнее персональная консультация.*
    """)
    st.markdown(
        "<p style='font-size:11px; color:#aaaaaa;'>Данный сервис носит исключительно информационный характер и не является официальной консультацией. Результаты подбора основаны на статистических данных прошлых лет и не гарантируют поступление. Приёмная кампания зависит от множества факторов которые невозможно предсказать заранее — статистика прошлых лет обычно хорошо отражает реальность, но никто не застрахован от неожиданных скачков конкурса и изменения проходных баллов. Сервис не несёт ответственности за решения принятые на основе предоставленной информации.</p>",
        unsafe_allow_html=True
    )
def send_email(to_email, result_df, search_params):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        CHANCE_COLORS = {
            "Уверенно": "D6EED2", "Реалистично": "D0E8F5", "Вероятно": "FFF0D6",
            "Рискованно": "FAE0E0", "Маловероятно": "EBEBEB",
            "Квоты и БВИ": "D7EAF3", "Общего конкурса не было": "D8D8D8",
            "Нет данных": "F5F5F5", "Нет оценки — не указан балл за ДВИ": "F5F5F5",
        }
        CHANCE_TEXT = {
            "Уверенно": "1E6B14", "Реалистично": "0D5A8A", "Вероятно": "8A5A00",
            "Рискованно": "8A1A1A", "Маловероятно": "555555",
            "Квоты и БВИ": "1A5C7A", "Общего конкурса не было": "4A4A4A",
            "Нет данных": "888888", "Нет оценки — не указан балл за ДВИ": "888888",
        }
        HEADER_BG = "379FFC"
        HEADER_FG = "FFFFFF"
        ROW_ODD   = "FFFFFF"
        ROW_EVEN  = "F4F7FB"
        ACCENT    = "379FFC"
        COL_W     = 90

        dashed = Side(style='dashDot', color='8BB8E8')
        thin_g = Side(style='thin', color='DDDDDD')
        no_s   = Side(style=None)

        def get_border(ci, bottom=True):
            b = thin_g if bottom else no_s
            l = dashed if ci == 6 else no_s
            r = dashed if ci in [1, 6, 8] else no_s
            return Border(bottom=b, left=l, right=r)

        def calc_height(text, col_width=90, font_size=9, base_height=15):
            if not text: return base_height
            chars_per_line = int(col_width * 1.15)
            lines = max(1, -(-len(str(text)) // chars_per_line))
            return max(base_height, lines * (font_size * 1.8))

        wb = Workbook()

        # ── Лист 1: Важная информация ──
        ws3 = wb.active
        ws3.title = "Важная информация"
        ws3.column_dimensions['A'].width = COL_W

        rows_info = [
            ('📋 Результаты вашего запроса — во вкладке "Результаты"', "title"),
            ("Ниже — важная информация о том, как читать таблицу и расставлять приоритеты.", "subtitle"),
            ("", "gap"),
            ("О СЕРВИСЕ", "header"),
            ("Данные актуальны на приёмную кампанию 2026 года.", "text"),
            ("Результаты основаны на статистике прошлого года и носят рекомендательный характер.", "text"),
            ("Таблица только для поступающих на общих основаниях на основном этапе.", "text"),
            ("Квотники и олимпиадники — для вас нужна персональная консультация.", "text"),
            ("", "gap"),
            ("РАСШИФРОВКА ШАНСОВ", "header"),
            ("Уверенно — ваш балл выше среднего на 10+. Это подстраховочный вариант на который вы точно пройдёте. Рек. приоритет: 3–5+", "D6EED2"),
            ("Реалистично — ваш балл выше среднего. Хорошие шансы. Рек. приоритет: 2–3", "D0E8F5"),
            ("Вероятно — ваш балл ниже среднего, но выше проходного на 5+. Рек. приоритет: 1–2", "FFF0D6"),
            ("Рискованно — ваш балл близко к проходному (не более чем на 5 выше или 15 ниже). Ставьте приоритет 1 только если очень хочется и есть подстраховка.", "FAE0E0"),
            ("Маловероятно — ваш балл ниже проходного на 15+. Шансы крайне малы.", "EBEBEB"),
            ("Квоты и БВИ — несмотря на наличие бюджетных мест, на данный момент они выделены только под квоты. Если желающих по квотам на эти специальности не будет, места перейдут олимпиадникам, а затем обычным поступающим на основном этапе. Это станет известно 3 августа.", "D7EAF3"),
            ("Общего конкурса не было — в прошлом году зачисление на основном этапе не проводилось, были зачислены только абитуриенты по квотам и БВИ (олимпиадники).", "D8D8D8"),
            ("Нет данных — новая специальность, статистики нет.", "F5F5F5"),
            ("", "gap"),
            ("ПРО ПРИОРИТЕТЫ", "header"),
            ("В одном вузе можно выбрать не более 5 кодов специальностей, но количество профилей не ограничено — поэтому приоритетов тоже может быть больше 5.", "text"),
            ("1–2 приоритет: амбициозные варианты.", "text"),
            ("3–4 приоритет: реалистичные, стабильные.", "text"),
            ("4+ приоритет: уверенное зачисление.", "text"),
            ("", "gap"),
            ("ВАЖНО: СОГЛАСИЕ НА ЗАЧИСЛЕНИЕ", "header"),
            ("Зачисление не происходит без подачи согласия.", "text"),
            ("Подать до 12:00 (мск) 5 августа в один из вузов — только один одновременно.", "text"),
            ("", "gap"),
            ("СРОКИ", "header"),
            ("До 25 июля — подача документов в вузы.", "text"),
            ("Срок для ДВИ короче — даты и формат уточняйте на сайте вуза.", "text"),
            ("27 июля — публикация конкурсных списков.", "text"),
            ("5 августа 12:00 (мск) — последний срок подачи согласия на зачисление.", "text"),
            ("", "gap"),
            ("ИНДИВИДУАЛЬНЫЕ ДОСТИЖЕНИЯ", "header"),
            ("Если есть достижения не учтённые в таблице — прибавьте самостоятельно. Суммарно не более 10 баллов. Полный список на сайте вуза.", "text"),
            ("", "gap"),
            ("СЛЕДУЮЩИЙ ШАГ", "header"),
            ("Отслеживайте позицию в конкурсных списках на сайтах вузов или на Госуслугах.", "text"),
            ("Не бойтесь раздутых списков — не все абитуриенты в них ваши реальные конкуренты.", "text"),
            ("", "gap"),
            ("Данный сервис носит исключительно информационный характер и не является официальной консультацией. Результаты не гарантируют поступление. Сервис не несёт ответственности за решения принятые на основе предоставленной информации.", "legal"),
            ("", "gap"),
            ("Для записи на консультацию пишите в телеграм-аккаунт @vuzline_webinar\nВопросы и предложения по таблице: result@vuzline.ru", "contact"),
        ]

        for rn, (text, kind) in enumerate(rows_info, 1):
            cell = ws3.cell(row=rn, column=1, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if kind == "title":
                cell.font = Font(name="Montserrat", size=16, bold=True, color=HEADER_BG)
                ws3.row_dimensions[rn].height = 42
            elif kind == "subtitle":
                cell.font = Font(name="Montserrat", size=10, color="555555")
                ws3.row_dimensions[rn].height = 22
            elif kind == "gap":
                ws3.row_dimensions[rn].height = 10
            elif kind == "header":
                cell.font = Font(name="Montserrat", size=10, bold=True, color=HEADER_BG)
                cell.fill = PatternFill("solid", fgColor="EBF4FF")
                ws3.row_dimensions[rn].height = 24
            elif kind == "legal":
                cell.font = Font(name="Montserrat", size=8, color="AAAAAA")
                ws3.row_dimensions[rn].height = calc_height(text, COL_W, 8)
            elif kind == "contact":
                cell.font = Font(name="Montserrat", size=9, bold=True, color=ACCENT)
                ws3.row_dimensions[rn].height = 36
            elif len(kind) == 6:
                cell.font = Font(name="Montserrat", size=9, color="1A1A1A")
                cell.fill = PatternFill("solid", fgColor=kind)
                ws3.row_dimensions[rn].height = calc_height(text, COL_W, 9)
            else:
                cell.font = Font(name="Montserrat", size=9, color="1A1A1A")
                ws3.row_dimensions[rn].height = calc_height(text, COL_W, 9)

        # ── Лист 2: Результаты ──
        ws = wb.create_sheet("Результаты")
        df_out = result_df.copy()
        chance_label = {
            "podstrahovka": "Уверенно", "realistic": "Реалистично",
            "probable": "Вероятно", "risky": "Рискованно",
            "unlikely": "Маловероятно",
            "quota_bvi": "Квоты и БВИ",
            "no_competition": "Общего конкурса не было",
            "new": "Нет данных",
            "no_dvi_score": "Нет оценки — не указан балл за ДВИ",
        }
        if "Шансы" in df_out.columns:
            df_out["Шансы"] = df_out["Шансы"].map(lambda x: chance_label.get(x, x))

        cols = list(df_out.columns)
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = Font(bold=True, color=HEADER_FG, name="Montserrat", size=9)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = get_border(ci, bottom=False)
        ws.row_dimensions[1].height = 32

        for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
            er = row_idx
            row = pd.Series(row._asdict())
            bg = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
            for ci, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=er, column=ci, value=str(value) if value is not None else "")
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = get_border(ci)
                if ci in [1, 2, 3, 4, 5]:
                    cell.font = Font(name="Montserrat", size=9, color="1A1A1A", bold=(ci == 1))
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                elif ci == 6:
                    cell.font = Font(name="Montserrat", size=9, color="1A1A1A", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 12:
                    c = CHANCE_COLORS.get(str(value), bg)
                    t = CHANCE_TEXT.get(str(value), "1A1A1A")
                    cell.fill = PatternFill("solid", fgColor=c)
                    cell.font = Font(name="Montserrat", size=9, color=t, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 13:
                    cell.font = Font(name="Montserrat", size=9, color=ACCENT, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = Font(name="Montserrat", size=9, color="1A1A1A")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        col_widths = {1:14,2:18,3:22,4:32,5:30,6:7,7:13,8:13,9:13,10:10,11:13,12:18,13:11,14:9,15:9,16:9,17:9}
        for ci, w in col_widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"

        # ── Лист 3: Запрос ──
        ws2 = wb.create_sheet("Запрос")
        params_df = pd.DataFrame([{k: str(v) for k, v in search_params.items()}])
        for ci, col_name in enumerate(params_df.columns, 1):
            cell = ws2.cell(row=1, column=ci, value=col_name)
            cell.font = Font(bold=True, color=HEADER_FG, name="Montserrat", size=9)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci, (col_name, value) in enumerate(params_df.iloc[0].items(), 1):
            cell = ws2.cell(row=2, column=ci, value=str(value) if value is not None else "")
            cell.font = Font(name="Montserrat", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws2.column_dimensions[get_column_letter(ci)].width = 28
        ws2.row_dimensions[1].height = 28
        ws2.row_dimensions[2].height = 40

        # Сохраняем в буфер
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Создаём письмо
        msg = MIMEMultipart()
        msg["From"] = st.secrets["EMAIL_FROM"]
        msg["To"] = to_email
        msg["Subject"] = "Ваша таблица подбора вузов — Vuzline"

        body = """
Здравствуйте!

Ваша персональная таблица подбора вузов готова. Она прикреплена к этому письму.

В таблице три листа:
- Важная информация — прочитайте сначала!
- Результаты — все подходящие специальности с оценкой шансов
- Запрос — параметры вашего поиска

Если у вас возникнут вопросы или вы хотите разобрать результаты подробнее — запишитесь на персональную консультацию со скидкой 500 руб. по промокоду VUZLINE500.
Для записи пишите в телеграм-аккаунт @vuzline_webinar.

Удачи с поступлением!
Команда Vuzline
vuzline.ru
        """
        msg.attach(MIMEText(body, "plain", "utf-8"))

        attachment = MIMEBase("application", "octet-stream")
        attachment.set_payload(buf.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            "attachment; filename=\"vuzline_results.xlsx\""
        )
        msg.attach(attachment)

        try:
            with smtplib.SMTP_SSL("smtp.yandex.ru", 465) as server:
                server.login(
                    str(st.secrets["EMAIL_FROM"]),
                    str(st.secrets["EMAIL_PASSWORD"])
                )
                server.sendmail(
                    str(st.secrets["EMAIL_FROM"]),
                    str(to_email),
                    msg.as_string()
                )
        except Exception as smtp_err:
            import traceback
            st.error(f"SMTP ошибка: {smtp_err}")
            st.code(traceback.format_exc())
            raise

        return True
    except Exception as e:
        import traceback
        st.error(f"Ошибка отправки письма: {e}")
        st.code(traceback.format_exc())
        return False
def show_results(result, flow=1, paid=False, selected_areas=None):
    result_few = pd.DataFrame()
    result_backup = pd.DataFrame()
    if len(result) == 0:
        st.warning("По вашему запросу ничего не найдено.")
        return

    st.success(f"Найдено {len(result)} специальностей")
    result = result.copy()
    result["chance_order"] = result["Шансы"].map(CHANCE_ORDER)
    result = result.sort_values(["Город", "Вуз", "chance_order"]).drop("chance_order", axis=1)
    result["Шансы"] = result["Шансы"].map(CHANCE_LABEL)

    # Тоггл скрытия вариантов сильно ниже уровня абитуриента
    hide_low = st.toggle(
        "Скрыть специальности сильно ниже моего уровня",
        value=False,
        help="Скрывает строки где ваш балл превышает проходной на 70+ баллов"
    )
    if hide_low and "Конкурсный балл" in result.columns and "Проходной балл" in result.columns:
        pb_num = pd.to_numeric(result["Проходной балл"], errors="coerce")
        score_num = pd.to_numeric(result["Конкурсный балл"], errors="coerce")
        mask = (score_num - pb_num) < 70
        result = result[mask | pb_num.isna() | score_num.isna()]

    # Предупреждение когда много вариантов но шансы в основном плохие
    good_statuses_set = {"🟢 Уверенно", "🔵 Реалистично", "🟡 Вероятно"}
    good_count = result["Шансы"].isin(good_statuses_set).sum()
    if len(result) >= 10 and good_count == 0:
        st.warning("⚠️ Среди найденных вариантов нет специальностей с хорошими шансами — все результаты относятся к категориям «Рискованно», «Маловероятно» или «Нет данных». Рекомендуем снизить планку или расширить список вузов.")
    elif len(result) >= 10 and good_count / len(result) < 0.15:
        st.warning(f"⚠️ Среди найденных вариантов только {good_count} с хорошими шансами. Большинство — «Рискованно» или «Маловероятно». Рекомендуем добавить другие профессиональные области, чтобы увеличить количество вариантов, или уберите их вовсе.")

    if flow == 1:
        vuz_counts = result.groupby("Вуз")["Код и специальность"].nunique()
        overloaded = vuz_counts[vuz_counts > 5]
        if len(overloaded) > 0:
            for vuz, count in overloaded.items():
                st.warning(f"⚠️ {vuz}: найдено {count} кодов специальностей — при подаче документов выберите не более 5.")
        if len(result) < 10:
            st.info("💡 Найдено мало вариантов. Попробуйте расширить поиск — добавьте другие вузы или специальности, чтобы получить больше подходящих результатов.")

    if flow == 2:
        CHANCE_PRIORITY = {
            "🟡 Вероятно": 0, "🔵 Реалистично": 1, "🟢 Уверенно": 2,
            "🔴 Рискованно": 3, "⚫ Маловероятно": 4,
            "🔹 Квоты и БВИ": 5, "◾ Общего конкурса не было": 6,
            "⬜ Нет данных": 7, "⬜ Нет оценки — не указан балл за ДВИ": 8,
        }
        result["_chance_p"] = result["Шансы"].map(CHANCE_PRIORITY)
        result["_pb_num"] = pd.to_numeric(result["Проходной балл"], errors="coerce").fillna(0)
        result = result.sort_values(["Город", "Вуз", "_chance_p", "_pb_num"],
                                     ascending=[True, True, True, False])
        result = result.drop(columns=["_pb_num"])

        good_zones = {"🟡 Вероятно", "🔵 Реалистично", "🟢 Уверенно"}
        real_competition = result[
            ~((pd.to_numeric(result["Проходной балл"], errors="coerce") <= 1) &
              (pd.to_numeric(result["Средний балл"], errors="coerce") <= 1))
        ]
        vuz_good_count = real_competition[real_competition["Шансы"].isin(good_zones)].groupby("Вуз").size()
        # Если вузов с 3+ хорошими вариантами меньше 3 — снимаем отсечку и показываем все
        vuz_with_3plus = vuz_good_count[vuz_good_count >= 3]
        if len(vuz_with_3plus) < 3:
            min_good_options = 1
        else:
            min_good_options = 3
        main_vuz = vuz_good_count[vuz_good_count >= min_good_options].sort_values(ascending=False)
        few_vuz = vuz_good_count[(vuz_good_count >= 1) & (vuz_good_count < min_good_options)].sort_values(ascending=False)
        few_vuz_list = list(few_vuz.index)

        def sort_vuz_by_rating(vuz_list):
            rated = sorted(
                [v for v in vuz_list if get_vuz_rating(v) < 999 and main_vuz.get(v, 0) >= 4],
                key=lambda v: get_vuz_rating(v)
            )
            unrated = [v for v in vuz_list if v not in rated]
            return rated + unrated

        def build_vuz_block(vuz_list):
            rows = []
            for vuz in vuz_list:
                vuz_df = result[result["Вуз"] == vuz].copy()
                seen_codes = set()
                for _, row in vuz_df.iterrows():
                    code_prefix = row["Код и специальность"].split(" ")[0][:5]
                    if code_prefix not in seen_codes:
                        if len(seen_codes) >= 5:
                            continue
                        seen_codes.add(code_prefix)
                    rows.append(row)
            df_out = pd.DataFrame(rows).reset_index(drop=True) if rows else pd.DataFrame()
            if len(df_out) > 0 and "_chance_p" in df_out.columns:
                df_out = df_out.drop(columns=["_chance_p"])
            return df_out

        main_vuz_list = list(main_vuz.index)
        if selected_areas:
            area_prefixes = set()
            for area in selected_areas:
                for prefix in AREA_GROUPS.get(area, []):
                    area_prefixes.add(prefix)
            good_result = real_competition[real_competition["Шансы"].isin(good_zones)]
            def vuz_in_area(vuz):
                vuz_codes = good_result[good_result["Вуз"] == vuz]["Код и специальность"]
                return any(str(c).split(".")[0] in area_prefixes for c in vuz_codes)
            area_vuz_list = [v for v in main_vuz_list if vuz_in_area(v)]
            backup_vuz_list = [v for v in main_vuz_list if not vuz_in_area(v)]
            top_area = sort_vuz_by_rating(area_vuz_list)[:7]
            top_backup = sort_vuz_by_rating(backup_vuz_list)[:7]
        else:
            top_area = sort_vuz_by_rating(main_vuz_list)[:7]
            top_backup = []

        result_main = build_vuz_block(top_area)
        result_backup = build_vuz_block(top_backup) if top_backup else pd.DataFrame()

        result_few_rows = []
        for vuz in few_vuz_list:
            vuz_df = result[(result["Вуз"] == vuz) & (result["Шансы"].isin(good_zones))].copy()
            seen_codes = set()
            for _, row in vuz_df.iterrows():
                code_prefix = row["Код и специальность"].split(" ")[0][:5]
                if code_prefix not in seen_codes:
                    if len(seen_codes) >= 5:
                        continue
                    seen_codes.add(code_prefix)
                result_few_rows.append(row)
        result_few = pd.DataFrame(result_few_rows).reset_index(drop=True) if result_few_rows else pd.DataFrame()
        if len(result_few) > 0 and "_chance_p" in result_few.columns:
            result_few = result_few.drop(columns=["_chance_p"])

        result = result_main if len(result_main) > 0 else pd.DataFrame()

        if selected_areas:
            if len(result_main) > 0:
                st.subheader("🎯 Ваши направления")
                st.caption(f"Топ-{len(top_area)} вузов по выбранным областям")
            else:
                st.warning("По выбранным направлениям не нашлось вузов с хорошими шансами. Ниже — подстраховочные варианты из других областей.")
        else:
            st.info(f"Показаны топ-{len(top_area)} вузов с наибольшим количеством подходящих специальностей")
        if len(result_few) > 0:
            st.caption(f"Ещё {len(few_vuz_list)} вузов с 1-2 подходящими вариантами показаны ниже")

    if not paid:
        preview_cols = ["Город", "Вуз", "Факультет", "Код и специальность", "Профиль"]
        preview = result[[c for c in preview_cols if c in result.columns]].copy()
        st.dataframe(preview, use_container_width=True, hide_index=True)
        st.info("""
🔒 **Полная таблица доступна после оплаты**

В полной версии вы увидите:
- Проходной и средний балл
- Ваш конкурсный балл
- Оценку шансов и рекомендуемый приоритет
- Баллы за индивидуальные достижения
- Возможность скачать таблицу в Excel

**Стоимость: 2 490 руб.**
        """)
        if st.button("💳 Оплатить и получить полную таблицу", type="primary", key="pay_btn"):
            if not st.session_state.get("user_email"):
                st.error("Введите email для получения таблицы")
            else:
                try:
                    # Генерируем order_id заранее и передаём в return_url
                    order_id = str(uuid.uuid4())
                    st.session_state["pending_order_id"] = order_id
                    return_url = f"https://vuzline-2026.streamlit.app/?order_id={order_id}"
                    payment_url, payment_id = create_payment(
                        amount=2490,
                        description="Подбор вузов по ЕГЭ — полная таблица",
                        return_url=return_url,
                        order_id=order_id
                    )
                    # Сохраняем данные в файл
                    search_params = {
                        "Дата и время": datetime.now().strftime("%d.%m.%Y %H:%M"),
                        "Email": st.session_state.get("user_email", ""),
                        "Предметы и баллы": str(st.session_state.get("last_subjects", {})),
                        "Города": str(st.session_state.get("last_cities", [])),
                        "ГТО": str(st.session_state.get("last_gto", "Нет")),
                        "Аттестат": str(st.session_state.get("last_attestat", False)),
                        "Балл за ДВИ": str(st.session_state.get("last_dvi", "")),
                        "Payment ID": str(payment_id),
                    }
                    result_df = pd.DataFrame.from_dict(st.session_state["last_result"])
                    save_payment_data(
                        order_id, result_df, search_params,
                        st.session_state["user_email"],
                        st.session_state.get("last_flow", 2),
                        payment_id=payment_id
                    )
                    st.session_state["payment_id"] = payment_id
                    st.session_state["payment_url"] = payment_url
                    st.rerun()
                except Exception as e:
                    st.error(f"Ошибка при создании платежа: {e}")
    else:
        st.dataframe(result, use_container_width=True, hide_index=True)
        show_disclaimers()
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
            result.to_excel(writer, index=False, sheet_name="Результаты")
            workbook = writer.book
            worksheet = writer.sheets["Результаты"]
            num_fmt = workbook.add_format({"num_format": "0"})
            for col_name in ["Мест", "Проходной балл", "Средний балл",
                              "Ваш балл (ЕГЭ)", "Конкурсный балл",
                              "ГТО золото", "ГТО серебро", "ГТО бронза", "Аттестат"]:
                if col_name in result.columns:
                    col_idx = result.columns.get_loc(col_name)
                    worksheet.set_column(col_idx, col_idx, 12, num_fmt)
        st.download_button(
            "📥 Скачать таблицу Excel",
            data=buf.getvalue(),
            file_name="результаты_подбора.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("""
💡 **Хотите разобрать результаты вместе?**
Запишитесь на персональную консультацию со скидкой 500 руб. по промокоду **VUZLINE500**
        """)

    if flow == 2 and len(result_backup) > 0:
        st.divider()
        st.subheader("🛡 Подстраховочные варианты")
        st.caption("Эти вузы не относятся к выбранным направлениям, но там есть специальности с хорошими шансами — стоит рассмотреть как запасной вариант")
        if not paid:
            preview_cols = ["Город", "Вуз", "Факультет", "Код и специальность", "Профиль"]
            preview_backup = result_backup[[c for c in preview_cols if c in result_backup.columns]].copy()
            st.dataframe(preview_backup, use_container_width=True, hide_index=True)
        else:
            st.dataframe(result_backup, use_container_width=True, hide_index=True)

    if flow == 2 and len(result_few) > 0:
        with st.expander("📋 Вузы с 1-2 подходящими вариантами"):
            st.caption("В этих вузах мало подходящих специальностей под ваши предметы и баллы")
            if not paid:
                preview_cols = ["Город", "Вуз", "Факультет", "Код и специальность", "Профиль"]
                preview_few = result_few[[c for c in preview_cols if c in result_few.columns]].copy()
                st.dataframe(preview_few, use_container_width=True, hide_index=True)
            else:
                st.dataframe(result_few, use_container_width=True, hide_index=True)


# ─── ИНТЕРФЕЙС ────────────────────────────────────────────────────────────
df = load_data()
Configuration.account_id = st.secrets["YUKASSA_SHOP_ID"]
Configuration.secret_key = st.secrets["YUKASSA_SECRET_KEY"]
city_options = get_city_options(df)
all_codes = get_all_codes(df)

# Проверяем payment_id из URL и отправляем письмо
query_params = st.query_params
order_id_from_url = query_params.get("order_id", "")

if order_id_from_url and not st.session_state.get(f"sent_{order_id_from_url}"):
    if get_email_sent_status(order_id_from_url):
        # Письмо уже отправлено вебхуком — просто сообщаем пользователю
        st.session_state[f"sent_{order_id_from_url}"] = True
        st.success("✅ Таблица уже отправлена вам на почту!")
    else:
        data = load_payment_data(order_id_from_url)
        if data:
            payment_id = data.get("payment_id", "")
            if check_payment_status(payment_id):
                result_df = pd.DataFrame.from_dict(data["result"])
                result_df = result_df.astype(str).replace('None', '').replace('nan', '')
                if send_email(data["user_email"], result_df, data["search_params"]):
                    st.session_state[f"sent_{order_id_from_url}"] = True
                    mark_email_sent(order_id_from_url)
                    st.success(f"✅ Таблица отправлена на {data['user_email']}!")
                else:
                    st.error("Ошибка отправки письма. Напишите нам на result@vuzline.ru и мы пришлём таблицу вручную.")
        else:
            st.info("Платёж обрабатывается... Обновите страницу через минуту.")

st.title("🎓 Подбор вузов по ЕГЭ")

st.info("""
⚠️ **Важно перед использованием**

Этот сервис подходит только для поступающих **на бюджет на общих основаниях на основном этапе** приёмной кампании 2026.

Если вы поступаете по квоте, как победитель олимпиады или с другими особыми условиями — данные прогнозы вам не подойдут. С такими кейсами приходите на персональную консультацию.
""")

# Если есть payment_url — показываем кнопку перехода к оплате
if "payment_url" in st.session_state:
    st.link_button("💳 Перейти к оплате (2 490 руб.)", st.session_state["payment_url"], type="primary")
    st.caption("После оплаты таблица придёт на почту автоматически в течение пары минут. Также вы можете вернуться на эту страницу, чтобы открыть её здесь.")

    if st.button("❌ Отменить и начать заново"):
        del st.session_state["payment_url"]
        if "payment_id" in st.session_state:
            del st.session_state["payment_id"]
        st.rerun()
    st.stop()

flow = st.radio(
    "Как вы хотите искать?",
    ["🔍 Подобрать варианты по моим ЕГЭ",
     "🎯 Я знаю вузы и специальности которые хочу"],
    horizontal=True
)

st.divider()

st.subheader("Введите баллы ЕГЭ")
subjects = {}

rus = st.number_input("Русский язык *", min_value=0, max_value=100,
                       value=None, placeholder="Введите балл")
if rus and rus > 0:
    subjects["Русский язык"] = rus

subjects_list = ["Математика","Обществознание","История","Иностранный язык",
                 "Биология","Химия","Физика","Информатика","География","Литература"]
selected_subj = st.multiselect("Выберите остальные предметы *", subjects_list)

if selected_subj:
    cols = st.columns(min(len(selected_subj), 3))
    for i, subj in enumerate(selected_subj):
        with cols[i % 3]:
            score = st.number_input(subj, min_value=0, max_value=100,
                                     value=None, placeholder="Введите балл", key=f"score_{subj}")
            if score and score > 0:
                subjects[subj] = score

st.subheader("Индивидуальные достижения")
st.caption("Суммарно не более 10 баллов")
col1, col2 = st.columns(2)
with col1:
    gto = st.selectbox("ГТО", ["Нет", "Золото", "Серебро", "Бронза"])
with col2:
    attestat = st.checkbox("Аттестат с отличием")

dvi_score = st.number_input(
    "Балл за ДВИ (если уже известен)",
    min_value=0, max_value=1000,
    value=None, placeholder="Необязательно",
    help="ДВИ — дополнительное вступительное испытание в вузе. Если несколько ДВИ — введите суммарный балл."
)

gto_val = gto if gto != "Нет" else None
st.divider()
st.subheader("Введите email для получения таблицы")
user_email = st.text_input(
    "Email *",
    placeholder="example@mail.ru",
    help="На этот адрес мы отправим полную таблицу после оплаты"
)
if user_email:
    st.session_state["user_email"] = user_email

st.divider()

# ─── ФЛОУ 2 ───────────────────────────────────────────────────────────────
if flow == "🔍 Подобрать варианты по моим ЕГЭ":
    st.subheader("Выберите города (до 3)")
    selected_cities = st.multiselect(
        "Города поиска *", city_options, max_selections=3,
        help="Москва и МО / Питер и ЛО идут как один выбор"
    )
    show_dvi = st.toggle(
        "Показывать специальности с ДВИ", value=False,
        help="ДВИ — дополнительное вступительное испытание в вузе"
    )
    selected_areas = st.multiselect(
        "Профессиональные области (необязательно)",
        list(AREA_GROUPS.keys()),
        help="Выберите интересующие направления — они будут показаны в первую очередь. Если не выбрать — покажем все подходящие специальности"
    )
    if st.button("🔍 Найти специальности", type="primary"):
        errors = []
        if not subjects.get("Русский язык"): errors.append("Введите балл за Русский язык")
        if len(subjects) < 2: errors.append("Введите минимум 2 предмета")
        if not selected_cities: errors.append("Выберите хотя бы один город")
        if errors:
            for e in errors: st.error(e)
        else:
            with st.spinner("Подбираем варианты..."):
                result = filter_rows_flow2(df, subjects, show_dvi, selected_cities, gto_val, attestat, dvi_score)
            if len(result) == 0:
                st.warning("По вашему запросу ничего не найдено.")
            else:
                st.session_state["last_result"] = result.to_dict()
                st.session_state["last_flow"] = 2
                st.session_state["last_areas"] = selected_areas
                st.session_state["last_subjects"] = subjects
                st.session_state["last_cities"] = selected_cities
                st.session_state["last_gto"] = gto_val
                st.session_state["last_attestat"] = attestat
                st.session_state["last_dvi"] = dvi_score
                st.rerun()

# ─── ФЛОУ 1 ───────────────────────────────────────────────────────────────
else:
    st.subheader("Выберите города (до 3)")
    selected_cities_flow1 = st.multiselect(
        "Города *", city_options, max_selections=3,
        help="Москва и МО / Питер и ЛО идут как один выбор",
        key="cities_flow1"
    )
    selected_areas_flow1 = st.multiselect(
        "Профессиональные области (необязательно)",
        list(AREA_GROUPS.keys()),
        help="Выберите интересующие направления — они сузят список кодов специальностей ниже",
        key="areas_flow1"
    )
    area_prefixes_flow1 = set()
    if selected_areas_flow1:
        for area in selected_areas_flow1:
            for prefix in AREA_GROUPS.get(area, []):
                area_prefixes_flow1.add(prefix)

    st.subheader("Выберите коды специальностей (необязательно)")
    if len(subjects) >= 2:
        available_codes = []
        seen = set()
        for _, row in df.iterrows():
            if selected_cities_flow1:
                city_raw = str(row.iloc[22]).strip()
                if get_city_group(city_raw) not in selected_cities_flow1: continue
            status = check_row(row, subjects)
            if status is not None:
                code = clean_str(row.iloc[25])
                if code and code not in seen:
                    if area_prefixes_flow1:
                        code_prefix = code.split('.')[0]
                        if code_prefix not in area_prefixes_flow1:
                            continue
                    seen.add(code)
                    available_codes.append(code)
        available_codes = sorted(available_codes)
        if selected_areas_flow1:
            st.caption(f"Показаны специальности по выбранным областям ({len(available_codes)} из {len(all_codes)})")
        else:
            st.caption(f"Показаны специальности подходящие под ваши предметы ({len(available_codes)} из {len(all_codes)})")
    else:
        available_codes = all_codes
        st.caption("Введите предметы ЕГЭ выше чтобы отфильтровать подходящие специальности")

    selected_codes = st.multiselect(
        "Коды специальностей *", available_codes,
        help="Начните вводить название или код для поиска"
    )
    st.subheader("Выберите вузы (до 5)")
    if selected_cities_flow1:
        vuz_options = []
        for city_group in selected_cities_flow1:
            vuz_options.extend(get_vuz_by_city(df, city_group))
        vuz_options = sorted(set(vuz_options))
        if selected_codes:
            expanded_for_ui = expand_code_set(selected_codes, df)
            filtered_vuz = set()
            for _, row in df.iterrows():
                city_raw = str(row.iloc[22]).strip()
                if get_city_group(city_raw) not in selected_cities_flow1: continue
                code = clean_str(row.iloc[25])
                if code in expanded_for_ui:
                    vuz = clean_str(row.iloc[23])
                    if vuz: filtered_vuz.add(vuz)
            vuz_options = sorted(filtered_vuz)
            st.caption(f"Показаны вузы где есть выбранные специальности ({len(vuz_options)})")
    else:
        vuz_options = []

    selected_vuz = st.multiselect("Вузы *", vuz_options, max_selections=5)

    if st.button("🎯 Найти по моему списку", type="primary"):
        errors = []
        if not subjects.get("Русский язык"): errors.append("Введите балл за Русский язык")
        if len(subjects) < 2: errors.append("Введите минимум 2 предмета")
        if not selected_vuz and not selected_codes: errors.append("Выберите хотя бы один вуз или один код специальности")
        if errors:
            for e in errors: st.error(e)
        else:
            with st.spinner("Ищем по вашему списку..."):
                result = filter_rows_flow1(df, subjects, selected_vuz, selected_codes,
                                           gto_val, attestat, selected_cities_flow1, dvi_score)
            if len(result) == 0:
                st.warning("По выбранным параметрам ничего не найдено. Попробуйте расширить поиск.")
            else:
                st.session_state["last_result"] = result.to_dict()
                st.session_state["last_flow"] = 1
                st.session_state["last_subjects"] = subjects
                st.session_state["last_cities"] = selected_cities_flow1
                st.session_state["last_gto"] = gto_val
                st.session_state["last_attestat"] = attestat
                st.session_state["last_dvi"] = dvi_score
                st.rerun()
# После rerun показываем результаты и кнопку оплаты
if "last_result" in st.session_state and "last_flow" in st.session_state:
    if "payment_url" not in st.session_state:
        result = pd.DataFrame.from_dict(st.session_state["last_result"])
        paid = st.session_state.get("paid", False)
        show_results(result, flow=st.session_state["last_flow"], paid=paid,
                     selected_areas=st.session_state.get("last_areas", []))
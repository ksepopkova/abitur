"""
Webhook-сервис для приёма HTTP-уведомлений от ЮКассы.

Когда платёж в ЮКассе становится "succeeded", ЮКасса присылает сюда
POST-запрос. Сервис находит соответствующие данные платежа в Google Sheets
(сохранённые туда приложением Streamlit при создании платежа),
формирует Excel-файл с результатами и отправляет письмо пользователю —
независимо от того, вернулся ли пользователь на сайт.

Эндпоинты:
  POST /webhook   — приём уведомлений от ЮКассы
  GET  /health     — проверка живости сервиса (для Render health-check)
"""

import os
import json
import gzip
import base64
import smtplib
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd
from fastapi import FastAPI, Request, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

import gspread
from google.oauth2.service_account import Credentials
from yookassa import Configuration, Payment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("webhook")

app = FastAPI()

# ── Конфигурация из переменных окружения ──
YUKASSA_SHOP_ID = os.environ["YUKASSA_SHOP_ID"]
YUKASSA_SECRET_KEY = os.environ["YUKASSA_SECRET_KEY"]
SHEETS_ID = os.environ["SHEETS_ID"]
GCP_SERVICE_ACCOUNT_JSON = os.environ["GCP_SERVICE_ACCOUNT_JSON"]  # весь JSON ключа сервисного аккаунта одной строкой
EMAIL_FROM = os.environ["EMAIL_FROM"]
EMAIL_PASSWORD = os.environ["EMAIL_PASSWORD"]

Configuration.account_id = YUKASSA_SHOP_ID
Configuration.secret_key = YUKASSA_SECRET_KEY


def get_sheets_client():
    creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON)
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive"]
    )
    return gspread.authorize(creds)


def find_row_by_order_id(sheet, order_id):
    cell = sheet.find(order_id, in_column=1)
    return cell.row if cell else None


def load_payment_row(order_id):
    client = get_sheets_client()
    sheet = client.open_by_key(SHEETS_ID).sheet1
    row_num = find_row_by_order_id(sheet, order_id)
    if not row_num:
        return None
    row_values = sheet.row_values(row_num)
    # Структура строки: order_id, payment_id, email, flow, datetime, status, compressed_data
    if len(row_values) < 7:
        return None
    status = row_values[5] if len(row_values) > 5 else ""
    compressed_b64 = row_values[6]
    raw_json = gzip.decompress(base64.b64decode(compressed_b64)).decode("utf-8")
    data = json.loads(raw_json)
    return {
        "row_num": row_num,
        "status": status,
        "user_email": data["user_email"],
        "flow": data["flow"],
        "search_params": data["search_params"],
        "result": data["result"],
    }


def mark_sent(order_id, row_num):
    client = get_sheets_client()
    sheet = client.open_by_key(SHEETS_ID).sheet1
    sheet.update_cell(row_num, 6, "sent")


# ── Логика формирования Excel и отправки письма ──
# (идентична send_email() в app.py — переиспользуется здесь напрямую,
# так как это отдельный, независимый сервис)

def build_excel(result_df, search_params):
    CHANCE_COLORS = {
        "Уверенно": "D6EED2", "Реалистично": "D0E8F5", "Вероятно": "FFF0D6",
        "Рискованно": "FAE0E0", "Маловероятно": "EBEBEB",
        "Квоты и БВИ": "D7EAF3", "Общего конкурса не было": "D8D8D8",
        "Нет данных": "F5F5F5", "Нет оценки — не указан балл за ДВИ": "F5F5F5",
    }
    CHANCE_TEXT = {
        "Уверенно": "1E6B14", "Реалистично": "0D5A8A", "Вероятно": "8A5A00",
        "Рискованно": "8A1A1A", "Маловероятно": "555555",
        "Квоты и БВИ": "1A5C7A", "Общего конкурса не было": "4A4A4A",
        "Нет данных": "888888", "Нет оценки — не указан балл за ДВИ": "888888",
    }
    HEADER_BG = "379FFC"
    HEADER_FG = "FFFFFF"
    ROW_ODD = "FFFFFF"
    ROW_EVEN = "F4F7FB"
    ACCENT = "379FFC"
    COL_W = 90

    dashed = Side(style='dashDot', color='8BB8E8')
    thin_g = Side(style='thin', color='DDDDDD')
    no_s = Side(style=None)

    def get_border(ci, bottom=True):
        b = thin_g if bottom else no_s
        l = dashed if ci == 6 else no_s
        r = dashed if ci in [1, 6, 8] else no_s
        return Border(bottom=b, left=l, right=r)

    def calc_height(text, col_width=90, font_size=9, base_height=15):
        if not text:
            return base_height
        chars_per_line = int(col_width * 1.15)
        lines = max(1, -(-len(str(text)) // chars_per_line))
        return max(base_height, lines * (font_size * 1.8))

    wb = Workbook()

    # ── Лист 1: Важная информация ──
    ws3 = wb.active
    ws3.title = "Важная информация"
    ws3.column_dimensions['A'].width = COL_W

    rows_info = [
        ('📋 Результаты вашего запроса — во вкладке "Результаты"', "title"),
        ("Ниже — важная информация о том, как читать таблицу и расставлять приоритеты.", "subtitle"),
        ("", "gap"),
        ("О СЕРВИСЕ", "header"),
        ("Данные актуальны на приёмную кампанию 2026 года.", "text"),
        ("Результаты основаны на статистике прошлого года и носят рекомендательный характер.", "text"),
        ("Таблица только для поступающих на общих основаниях на основном этапе.", "text"),
        ("Квотники и олимпиадники — для вас нужна персональная консультация.", "text"),
        ("", "gap"),
        ("РАСШИФРОВКА ШАНСОВ", "header"),
        ("Уверенно — ваш балл выше среднего на 10+. Это подстраховочный вариант на который вы точно пройдёте. Рек. приоритет: 3–5+", "D6EED2"),
        ("Реалистично — ваш балл выше среднего. Хорошие шансы. Рек. приоритет: 2–3", "D0E8F5"),
        ("Вероятно — ваш балл ниже среднего, но выше проходного на 5+. Рек. приоритет: 1–2", "FFF0D6"),
        ("Рискованно — ваш балл близко к проходному (не более чем на 5 выше или 15 ниже). Ставьте приоритет 1 только если очень хочется и есть подстраховка.", "FAE0E0"),
        ("Маловероятно — ваш балл ниже проходного на 15+. Шансы крайне малы.", "EBEBEB"),
        ("Квоты и БВИ — несмотря на наличие бюджетных мест, на данный момент они выделены только под квоты. Если желающих по квотам на эти специальности не будет, места перейдут олимпиадникам, а затем обычным поступающим на основном этапе. Это станет известно 3 августа.", "D7EAF3"),
        ("Общего конкурса не было — в прошлом году зачисление на основном этапе не проводилось, были зачислены только абитуриенты по квотам и БВИ (олимпиадники).", "D8D8D8"),
        ("Нет данных — новая специальность, статистики нет.", "F5F5F5"),
        ("", "gap"),
        ("ПРО ПРИОРИТЕТЫ", "header"),
        ("В одном вузе можно выбрать не более 5 кодов специальностей, но количество профилей не ограничено — поэтому приоритетов тоже может быть больше 5.", "text"),
        ("1–2 приоритет: амбициозные варианты.", "text"),
        ("3–4 приоритет: реалистичные, стабильные.", "text"),
        ("4+ приоритет: уверенное зачисление.", "text"),
        ("", "gap"),
        ("ВАЖНО: СОГЛАСИЕ НА ЗАЧИСЛЕНИЕ", "header"),
        ("Зачисление не происходит без подачи согласия.", "text"),
        ("Подать до 12:00 (мск) 5 августа в один из вузов — только один одновременно.", "text"),
        ("", "gap"),
        ("СРОКИ", "header"),
        ("До 25 июля — подача документов в вузы.", "text"),
        ("Срок для ДВИ короче — даты и формат уточняйте на сайте вуза.", "text"),
        ("27 июля — публикация конкурсных списков.", "text"),
        ("5 августа 12:00 (мск) — последний срок подачи согласия на зачисление.", "text"),
        ("", "gap"),
        ("ИНДИВИДУАЛЬНЫЕ ДОСТИЖЕНИЯ", "header"),
        ("Если есть достижения не учтённые в таблице — прибавьте самостоятельно. Суммарно не более 10 баллов. Полный список на сайте вуза.", "text"),
        ("", "gap"),
        ("СЛЕДУЮЩИЙ ШАГ", "header"),
        ("Отслеживайте позицию в конкурсных списках на сайтах вузов или на Госуслугах.", "text"),
        ("Не бойтесь раздутых списков — не все абитуриенты в них ваши реальные конкуренты.", "text"),
        ("", "gap"),
        ("Данный сервис носит исключительно информационный характер и не является официальной консультацией. Результаты не гарантируют поступление. Сервис не несёт ответственности за решения принятые на основе предоставленной информации.", "legal"),
        ("", "gap"),
        ("Для записи на консультацию пишите в телеграм-аккаунт @vuzline_webinar\nВопросы и предложения по таблице: result@vuzline.ru", "contact"),
    ]

    for rn, (text, kind) in enumerate(rows_info, 1):
        cell = ws3.cell(row=rn, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if kind == "title":
            cell.font = Font(name="Montserrat", size=16, bold=True, color=HEADER_BG)
            ws3.row_dimensions[rn].height = 42
        elif kind == "subtitle":
            cell.font = Font(name="Montserrat", size=10, color="555555")
            ws3.row_dimensions[rn].height = 22
        elif kind == "gap":
            ws3.row_dimensions[rn].height = 10
        elif kind == "header":
            cell.font = Font(name="Montserrat", size=10, bold=True, color=HEADER_BG)
            cell.fill = PatternFill("solid", fgColor="EBF4FF")
            ws3.row_dimensions[rn].height = 24
        elif kind == "legal":
            cell.font = Font(name="Montserrat", size=8, color="AAAAAA")
            ws3.row_dimensions[rn].height = calc_height(text, COL_W, 8)
        elif kind == "contact":
            cell.font = Font(name="Montserrat", size=9, bold=True, color=ACCENT)
            ws3.row_dimensions[rn].height = 36
        elif len(kind) == 6:
            cell.font = Font(name="Montserrat", size=9, color="1A1A1A")
            cell.fill = PatternFill("solid", fgColor=kind)
            ws3.row_dimensions[rn].height = calc_height(text, COL_W, 9)
        else:
            cell.font = Font(name="Montserrat", size=9, color="1A1A1A")
            ws3.row_dimensions[rn].height = calc_height(text, COL_W, 9)

    # ── Лист 2: Результаты ──
    ws = wb.create_sheet("Результаты")
    df_out = result_df.copy()
    chance_label = {
        "podstrahovka": "Уверенно", "realistic": "Реалистично",
        "probable": "Вероятно", "risky": "Рискованно",
        "unlikely": "Маловероятно",
        "quota_bvi": "Квоты и БВИ",
        "no_competition": "Общего конкурса не было",
        "new": "Нет данных",
        "no_dvi_score": "Нет оценки — не указан балл за ДВИ",
    }
    # Нормализуем Шансы: убираем emoji-префикс если данные уже обработаны
    emoji_strip = {
        "🟢 Уверенно": "podstrahovka",
        "🔵 Реалистично": "realistic",
        "🟡 Вероятно": "probable",
        "🔴 Рискованно": "risky",
        "⚫ Маловероятно": "unlikely",
        "🔹 Квоты и БВИ": "quota_bvi",
        "◾ Общего конкурса не было": "no_competition",
        "⬜ Нет данных": "new",
        "⬜ Нет оценки — не указан балл за ДВИ": "no_dvi_score",
    }
    if "Шансы" in df_out.columns:
        df_out.loc[:, "Шансы"] = df_out["Шансы"].map(
            lambda x: chance_label.get(emoji_strip.get(x, x), x)
        )

    cols = list(df_out.columns)
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = Font(bold=True, color=HEADER_FG, name="Montserrat", size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = get_border(ci, bottom=False)
    ws.row_dimensions[1].height = 32

    for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
        er = row_idx
        row = pd.Series(row._asdict())
        bg = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
        for ci, (col_name, value) in enumerate(row.items(), 1):
            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell_val = ""
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cell_val = value
            else:
                s = str(value)
                cell_val = "" if s in ("None", "nan") else s

            cell = ws.cell(row=er, column=ci, value=cell_val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = get_border(ci)
            if ci in [1, 2, 3, 4, 5]:
                cell.font = Font(name="Montserrat", size=9, color="1A1A1A", bold=(ci == 1))
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            elif ci == 6:
                cell.font = Font(name="Montserrat", size=9, color="1A1A1A", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 12:
                c = CHANCE_COLORS.get(str(value), bg)
                t = CHANCE_TEXT.get(str(value), "1A1A1A")
                cell.fill = PatternFill("solid", fgColor=c)
                cell.font = Font(name="Montserrat", size=9, color=t, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 13:
                cell.font = Font(name="Montserrat", size=9, color=ACCENT, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(name="Montserrat", size=9, color="1A1A1A")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = {1: 14, 2: 18, 3: 22, 4: 32, 5: 30, 6: 7, 7: 13, 8: 13, 9: 13, 10: 10, 11: 13, 12: 18, 13: 11, 14: 9, 15: 9, 16: 9, 17: 9}
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    # ── Лист 3: Запрос ──
    ws2 = wb.create_sheet("Запрос")
    params_df = pd.DataFrame([{k: str(v) for k, v in search_params.items()}])
    for ci, col_name in enumerate(params_df.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=col_name)
        cell.font = Font(bold=True, color=HEADER_FG, name="Montserrat", size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ci, (col_name, value) in enumerate(params_df.iloc[0].items(), 1):
        cell = ws2.cell(row=2, column=ci, value=str(value) if value is not None else "")
        cell.font = Font(name="Montserrat", size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws2.column_dimensions[get_column_letter(ci)].width = 28
    ws2.row_dimensions[1].height = 28
    ws2.row_dimensions[2].height = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def send_result_email(to_email, result_df, search_params):
    buf = build_excel(result_df, search_params)

    msg = MIMEMultipart()
    msg["From"] = EMAIL_FROM
    msg["To"] = to_email
    msg["Subject"] = "Ваша таблица подбора вузов — Vuzline"

    body = """
Здравствуйте!

Ваша персональная таблица подбора вузов готова. Она прикреплена к этому письму.

В таблице три листа:
- Важная информация — прочитайте сначала!
- Результаты — все подходящие специальности с оценкой шансов
- Запрос — параметры вашего поиска

Если у вас возникнут вопросы или вы хотите разобрать результаты подробнее — запишитесь на персональную консультацию со скидкой 500 руб. по промокоду VUZLINE500.
Для записи пишите в телеграм-аккаунт @vuzline_webinar.

Удачи с поступлением!
Команда Vuzline
vuzline.ru
    """
    msg.attach(MIMEText(body, "plain", "utf-8"))

    attachment = MIMEBase("application", "octet-stream")
    attachment.set_payload(buf.read())
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition",
        "attachment; filename=\"vuzline_results.xlsx\""
    )
    msg.attach(attachment)

    logger.info(f"Подключаемся к SMTP smtp.yandex.ru:587")
    with smtplib.SMTP("smtp.yandex.ru", 587, timeout=30) as server:
        server.ehlo()
        logger.info("ehlo выполнен")
        server.starttls()
        logger.info("starttls выполнен")
        server.login(EMAIL_FROM, EMAIL_PASSWORD)
        logger.info("login выполнен")
        server.sendmail(EMAIL_FROM, to_email, msg.as_string())
        logger.info("sendmail выполнен")


# ── Эндпоинты ──

@app.get("/health")
@app.head("/health")
async def health():
    return {"status": "ok"}


@app.post("/webhook")
async def yookassa_webhook(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    logger.info(f"Получено уведомление: {body.get('event')}")

    event = body.get("event")
    if event != "payment.succeeded":
        # Игнорируем другие типы событий (waiting_for_capture, canceled и т.д.)
        return {"status": "ignored", "event": event}

    payment_obj = body.get("object", {})
    payment_id = payment_obj.get("id")
    metadata = payment_obj.get("metadata", {})
    order_id = metadata.get("order_id")

    if not order_id:
        logger.warning("В уведомлении нет order_id в metadata")
        return {"status": "no_order_id"}

    # Доп. проверка статуса напрямую через API ЮКассы (защита от поддельных вебхуков)
    try:
        payment = Payment.find_one(payment_id)
        if payment.status != "succeeded":
            logger.warning(f"Платёж {payment_id} не succeeded по данным API: {payment.status}")
            return {"status": "not_succeeded"}
    except Exception as e:
        logger.error(f"Не удалось проверить платёж через API: {e}")
        raise HTTPException(status_code=500, detail="Payment verification failed")

    # Загружаем данные платежа из Google Sheets
    try:
        record = load_payment_row(order_id)
    except Exception as e:
        logger.error(f"Ошибка чтения данных из Google Sheets: {e}")
        raise HTTPException(status_code=500, detail="Failed to load payment data")

    if not record:
        logger.warning(f"Данные для order_id={order_id} не найдены в Google Sheets")
        return {"status": "data_not_found"}

    logger.info(f"Данные найдены, статус: '{record['status']}', email: {record['user_email']}")

    if record["status"] == "sent":
        logger.info(f"Письмо для order_id={order_id} уже было отправлено ранее")
        return {"status": "already_sent"}

    # Формируем и отправляем письмо
    try:
        logger.info(f"Загружены данные для order_id={order_id}, email={record['user_email']}")
        result_df = pd.DataFrame.from_dict(record["result"])
        # Очищаем None/nan не теряя числовые типы
        for col in result_df.columns:
            result_df[col] = result_df[col].apply(
                lambda x: "" if x is None or (isinstance(x, float) and pd.isna(x)) or str(x) in ("None", "nan") else x
            )
        logger.info(f"DataFrame сформирован, строк: {len(result_df)}")
        send_result_email(record["user_email"], result_df, record["search_params"])
        logger.info(f"Письмо успешно отправлено на {record['user_email']} (order_id={order_id})")
        mark_sent(order_id, record["row_num"])
        logger.info(f"Статус 'sent' записан в Google Sheets")
        return {"status": "sent"}
    except Exception as e:
        import traceback
        logger.error(f"Ошибка отправки письма: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to send email: {e}")
